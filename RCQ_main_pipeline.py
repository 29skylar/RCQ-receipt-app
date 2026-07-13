# RCQ_main_pipeline.py
import os
import re
import shutil
import pandas as pd
from datetime import datetime

from RCQ_config import INPUT_DIR, OUTPUT_EXCEL_PATH, PROCESSED_DIR, resolve_template_path
from RCQ_aws_engine import (
    extract_with_aws,
    extract_total_by_keyword,
    extract_total_from_text,
    extract_amount_by_currency,   # NEW
    detect_known_vendor,
    categorize_receipt,
)
from RCQ_gcp_engine import extract_with_gcp
from RCQ_pdf_handler import expand_pdf_to_pages

def safe_write(ws, row, column, value):
    """
    Writes a value to a cell, handling merged cells automatically.
    If the target is inside a merged range, writes to the top-left
    cell of that range instead (which is where merged data lives).
    """
    target_cell = ws.cell(row=row, column=column)
    coord = target_cell.coordinate

    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            # Redirect to the top-left cell of the merge
            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            top_left.value = value
            return top_left  # return the cell we actually wrote to

    # Not merged — write directly
    target_cell.value = value
    return target_cell

# ============================================================
# Helper functions
# ============================================================

def normalize_amount(amount_str):
    """
    Strips currency symbols and returns a real float (or None if invalid).
    Returning a number lets Excel apply numeric formatting.
    """
    if not amount_str:
        return None
    cleaned = re.sub(r'[^\d.,]', '', str(amount_str))
    cleaned = cleaned.replace(',', '')
    if cleaned.count('.') > 1:
        parts = cleaned.split('.')
        cleaned = ''.join(parts[:-1]) + '.' + parts[-1]
    try:
        return float(cleaned) if cleaned else None
    except ValueError:
        return None


def normalize_date(date_str):
    """
    Tries to parse many date formats and returns a real datetime object
    (or None if it can't be parsed). Excel will then format it as a real date.
    """
    if not date_str:
        return None

    cleaned = str(date_str).strip()

    # Handle Chinese date format: 2024年3月15日
    chinese_match = re.match(r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日?', cleaned)
    if chinese_match:
        y, m, d = chinese_match.groups()
        try:
            return datetime(int(y), int(m), int(d))
        except ValueError:
            pass

    cleaned = re.sub(r'\s+', ' ', cleaned)
    formats = [
        '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
        '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y',
        '%d/%m/%y', '%d-%m-%y',
        '%m/%d/%Y', '%m-%d-%Y',
        '%d %b %Y', '%d %B %Y', '%d-%b-%Y', '%d-%B-%Y',
        '%b %d %Y', '%B %d %Y',
        '%Y%m%d',
    ]
    for fmt in formats:
        try:
            return datetime.strptime(cleaned, fmt)
        except ValueError:
            continue

    return None


def normalize_store_name(name):
    """Trims whitespace, collapses newlines, removes trailing punctuation."""
    if not name:
        return ''
    name = re.sub(r'\s+', ' ', str(name)).strip()
    name = re.sub(r'[\*_\-=]+$', '', name).strip()
    return name


def pick_best_value(aws_val, gcp_val, prefer_longer=False):
    """Picks the best value when AWS and GCP both return something."""
    if aws_val and gcp_val:
        if prefer_longer and len(gcp_val) > len(aws_val) * 1.3:
            return gcp_val
        return aws_val
    return aws_val or gcp_val or ''


def compute_confidence_flag(total_final, aws_total, gcp_total):
    """
    Returns confidence based on whether we ended up with a total.
      HIGH    - we have a total AND both engines/methods agree
      REVIEW  - we have a total but methods disagree
      LOW     - no total found at all
    """
    if not total_final:
        return 'LOW'

    # Both APIs returned a total and they agree with our final pick
    if aws_total and gcp_total and aws_total == gcp_total == total_final:
        return 'HIGH'

    # Only one engine had it, but it matches the final pick
    if (aws_total == total_final and not gcp_total) or \
       (gcp_total == total_final and not aws_total):
        return 'HIGH'

    # We found something but engines disagreed or only keyword-search worked
    return 'REVIEW'


# ============================================================
# Main pipeline
# ============================================================

IMAGE_EXTS = ('.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.tif')
PDF_EXTS = ('.pdf',)


def _queue_receipt_file(filename, file_path, pdf_output_dir, files_to_process):
    """Add an image or PDF (expanded per page) to the processing queue."""
    lower = filename.lower()
    if lower.endswith(IMAGE_EXTS):
        files_to_process.append((filename, file_path))
    elif lower.endswith(PDF_EXTS):
        files_to_process.extend(expand_pdf_to_pages(filename, file_path, pdf_output_dir))


def collect_files_to_process(input_dir, pdf_temp_dir):
    """Scan a folder for receipt images/PDFs and return (display_name, path) pairs."""
    os.makedirs(pdf_temp_dir, exist_ok=True)
    files_to_process = []

    for filename in sorted(os.listdir(input_dir)):
        full_path = os.path.join(input_dir, filename)

        if not os.path.isfile(full_path):
            continue

        lower = filename.lower()

        if lower.endswith(IMAGE_EXTS) or lower.endswith(PDF_EXTS):
            _queue_receipt_file(filename, full_path, pdf_temp_dir, files_to_process)

    return files_to_process


def process_single_receipt(display_name, image_path):
    """Run AWS + GCP extraction on one receipt and return a result row dict."""
    print(f"\nProcessing: {display_name}")

    try:
        aws_data = extract_with_aws(image_path)
    except Exception as e:
        print(f"   AWS error: {e}")
        aws_data = {'store_name': '', 'date': '', 'time': '', 'total': '', 'subtotal': ''}

    try:
        gcp_data = extract_with_gcp(image_path)
    except Exception as e:
        print(f"   GCP error: {e}")
        gcp_data = {'store_name': '', 'date': '', 'time': '', 'total': '', 'subtotal': '', '_full_text': ''}

    store_name = pick_best_value(
        normalize_store_name(aws_data.get('store_name', '')),
        normalize_store_name(gcp_data.get('store_name', '')),
        prefer_longer=True
    )
    if not store_name:
        store_name = detect_known_vendor(gcp_data.get('_full_text', ''))
        if store_name:
            print(f"   Store detected by keyword: {store_name}")

    date_raw = pick_best_value(aws_data.get('date', ''), gcp_data.get('date', ''))
    date_final = normalize_date(date_raw)

    time_final = pick_best_value(
        aws_data.get('time', ''),
        gcp_data.get('time', '')
    ).strip()

    full_text_for_categorization = gcp_data.get('_full_text', '')
    category = categorize_receipt(store_name, full_text_for_categorization)
    print(f"   Category: {category}")

    aws_total = normalize_amount(aws_data.get('total', ''))
    gcp_total = normalize_amount(gcp_data.get('total', ''))

    keyword_total = ''
    gcp_text_total = ''
    currency_total = ''

    if aws_total and gcp_total and aws_total == gcp_total:
        total_final = aws_total
        print(f"   Both engines agreed on total: {total_final}")
    else:
        try:
            keyword_total = normalize_amount(extract_total_by_keyword(image_path))
        except Exception as e:
            print(f"   AWS keyword detection error: {e}")

        if not keyword_total:
            try:
                full_text = gcp_data.get('_full_text', '')
                gcp_text_total = normalize_amount(extract_total_from_text(full_text))
            except Exception as e:
                print(f"   GCP text scan error: {e}")

        if (not keyword_total and not gcp_text_total
                and not aws_total and not gcp_total):
            try:
                full_text = gcp_data.get('_full_text', '')
                currency_total = normalize_amount(extract_amount_by_currency(full_text))
            except Exception as e:
                print(f"   Currency detection error: {e}")

        total_final = (
            keyword_total
            or gcp_text_total
            or aws_total
            or gcp_total
            or currency_total
        )

    if total_final:
        print(f"   Total picked: {total_final} "
              f"(AWS_field={aws_total}, GCP_field={gcp_total}, "
              f"AWS_keyword={keyword_total}, GCP_text={gcp_text_total})")
    else:
        print("   No total could be detected for this receipt.")

    confidence = compute_confidence_flag(total_final, aws_total, gcp_total)

    return {
        'Filename': display_name,
        'Date': date_final,
        'Time': time_final,
        'Category': category,
        'Store_Name': store_name,
        'Amount': total_final,
        'Confidence_Flag': confidence,
    }


def build_excel_workbook(results):
    """Fill the expense template and return (output_path, excel_bytes, using_template)."""
    from io import BytesIO
    from openpyxl import Workbook, load_workbook

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    base, ext = os.path.splitext(OUTPUT_EXCEL_PATH)
    output_path = f"{base}_{timestamp}{ext}"

    template_path = resolve_template_path()
    using_template = os.path.exists(template_path)

    if using_template:
        print(f"\nLoading template from {template_path}...")
        wb = load_workbook(template_path, keep_links=False)
        ws = wb.active or wb.worksheets[0]
        start_row = 10
    else:
        print(f"\nTemplate not found at {template_path}; using simple export.")
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = "Receipts"
        headers = ['Item No.', 'Date', 'Supplier', 'Category', 'Amount', 'Confidence']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        start_row = 2

    for idx, row_data in enumerate(results):
        current_row = start_row + idx

        safe_write(ws, current_row, 1, idx + 1)

        date_cell = safe_write(ws, current_row, 2, row_data['Date'])
        if date_cell.value:
            date_cell.number_format = 'd-mmm-yyyy'

        safe_write(ws, current_row, 3, row_data['Store_Name'])
        safe_write(ws, current_row, 4, row_data['Category'])

        amount_col = 10 if using_template else 5
        confidence_col = 13 if using_template else 6

        amount_cell = safe_write(ws, current_row, amount_col, row_data['Amount'])
        if amount_cell.value is not None:
            amount_cell.number_format = '#,##0.00_);[Red](#,##0.00)'

        safe_write(ws, current_row, confidence_col, row_data['Confidence_Flag'])

    buffer = BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    wb.save(output_path)
    print(f"Successfully saved formatted results to: {output_path}\n")

    return output_path, excel_bytes, using_template


def summarize_results(results):
    """Return confidence counts for display."""
    return {
        'high': sum(1 for r in results if r['Confidence_Flag'] == 'HIGH'),
        'review': sum(1 for r in results if r['Confidence_Flag'] == 'REVIEW'),
        'low': sum(1 for r in results if r['Confidence_Flag'] == 'LOW'),
    }


def process_file_list(files_to_process):
    """Process receipts and return (results, output_path, excel_bytes)."""
    results = []
    for display_name, image_path in files_to_process:
        results.append(process_single_receipt(display_name, image_path))

    output_path, excel_bytes, used_template = build_excel_workbook(results)
    stats = summarize_results(results)
    print(f"   HIGH confidence:   {stats['high']}")
    print(f"   REVIEW needed:     {stats['review']}")
    print(f"   LOW confidence:    {stats['low']}")

    return results, output_path, excel_bytes, used_template


def process_uploaded_files(uploaded_files, work_dir):
    """
    Process Streamlit (or other) uploaded file objects saved under work_dir.
    uploaded_files: list of (original_filename, saved_path) tuples.
    Returns (results, output_path, excel_bytes, used_template).
    """
    pdf_temp_dir = os.path.join(work_dir, '_pdf_pages')
    files_to_process = []
    for filename, saved_path in uploaded_files:
        _queue_receipt_file(filename, saved_path, pdf_temp_dir, files_to_process)

    if not files_to_process:
        return [], None, None, False

    return process_file_list(files_to_process)


def process_receipts():
    """Main loop: process each receipt image (and PDF page) and save to Excel."""
    pdf_temp_dir = os.path.join(INPUT_DIR, '_pdf_pages')
    files_to_process = collect_files_to_process(INPUT_DIR, pdf_temp_dir)
    original_files_seen = set()

    print(f"\nTotal receipts to process: {len(files_to_process)}\n")

    results = []
    for display_name, image_path in files_to_process:
        result_row = process_single_receipt(display_name, image_path)
        results.append(result_row)
        original_filename = display_name.split(' (page ')[0] if ' (page ' in display_name else display_name
        original_files_seen.add(original_filename)

    output_path, _, _ = build_excel_workbook(results)
    stats = summarize_results(results)
    print(f"   HIGH confidence:   {stats['high']}")
    print(f"   REVIEW needed:     {stats['review']}")
    print(f"   LOW confidence:    {stats['low']}")

    # Move all processed files to the processed folder
    os.makedirs(PROCESSED_DIR, exist_ok=True)
    moved_count = 0
    for original_filename in original_files_seen:
        source_path = os.path.join(INPUT_DIR, original_filename)
        if not os.path.exists(source_path):
            continue

        destination_path = os.path.join(PROCESSED_DIR, original_filename)

        if os.path.exists(destination_path):
            base_name, ext = os.path.splitext(original_filename)
            timestamp_suffix = datetime.now().strftime('%Y%m%d_%H%M%S')
            destination_path = os.path.join(PROCESSED_DIR, f"{base_name}_{timestamp_suffix}{ext}")

        try:
            shutil.move(source_path, destination_path)
            moved_count += 1
        except Exception as e:
            print(f"Could not move {original_filename}: {e}")

    print(f"\nMoved {moved_count} file(s) to processed folder")

    if os.path.exists(pdf_temp_dir):
        try:
            shutil.rmtree(pdf_temp_dir)
            print("Cleaned up temporary PDF page images")
        except Exception as e:
            print(f"Could not clean up temp folder: {e}")


if __name__ == "__main__":
    process_receipts()